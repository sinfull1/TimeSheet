import openpyxl
import datetime
import pandas
import pdb
import pprint
from dIM import get_datetime_range
from openpyxl.compat import range as rangeEx


wb = openpyxl.load_workbook('template.xlsx')

sheetNames = wb.get_sheet_names()

empSheet = wb.get_sheet_by_name('Employees')
# target = wb.copy_worksheet(empSheet)

print (empSheet['A2'].value)
print (dir(empSheet))

pp = pprint.PrettyPrinter(indent=4)
pp.pprint(dir(wb))

# pp.pprint(dir(target))
# print (target.sheet_state)


# add dates 
currentDate = datetime.datetime.now()
currentMonth = currentDate.month
currentYear= currentDate.year
dObj,dMon = get_datetime_range(currentYear, currentMonth)

for col in range(27, 54):
     _ = empSheet.cell(column=col, row='A', value="{0}".format(get_column_letter(col)))
print(empSheet['AA10'].value)

empSheet.save('template.xlsx')

print (currentYear)